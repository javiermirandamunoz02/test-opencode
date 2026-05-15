from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Model
from django.http import HttpResponse, JsonResponse, HttpRequest
from django.utils import timezone
from datetime import date, timedelta
from .models import Nota, Tarea, EntregaTarea, Asistencia, Comentario
from apps.gestion.models import AsignacionCurso
from apps.core.decorators import rol_requerido


@login_required
def nota_list(request):
    user = request.user
    q = request.GET.get("q", "")
    tipo = request.GET.get("tipo", "")
    periodo = request.GET.get("periodo", "")

    if user.es_alumno():
        notas = Nota.objects.filter(alumno=user.perfil_alumno).select_related("asignatura", "profesor__usuario")
    elif user.es_profesor():
        notas = Nota.objects.filter(profesor=user.perfil_profesor).select_related("alumno__usuario", "asignatura")
    else:
        notas = Nota.objects.all().select_related("alumno__usuario", "asignatura", "profesor__usuario")

    if q:
        notas = notas.filter(
            Q(alumno__usuario__nombre__icontains=q) |
            Q(alumno__usuario__apellido__icontains=q) |
            Q(asignatura__nombre__icontains=q)
        )
    if tipo:
        notas = notas.filter(tipo=tipo)
    if periodo:
        notas = notas.filter(periodo__icontains=periodo)

    paginator = Paginator(notas, 15)
    page = paginator.get_page(request.GET.get("page"))

    return render(request, "academicos/nota_list.html", {
        "notas": page,
        "tipos": Nota.Tipo.choices,
        "q": q, "tipo_filtro": tipo, "periodo": periodo,
    })


@login_required
@rol_requerido("profesor", "admin")
def nota_create(request):
    user = request.user
    if request.method == "POST":
        Nota.objects.create(
            alumno_id=request.POST.get("alumno"),
            asignatura_id=request.POST.get("asignatura"),
            profesor=user.perfil_profesor if user.es_profesor() else None,
            valor=request.POST.get("valor"),
            tipo=request.POST.get("tipo", "examen"),
            periodo=request.POST.get("periodo", ""),
            observacion=request.POST.get("observacion", ""),
        )
        messages.success(request, "Nota cargada exitosamente")
        return redirect("nota_list")

    from apps.usuarios.models import Alumno
    from apps.gestion.models import Asignatura

    if user.es_profesor():
        asignaciones = AsignacionCurso.objects.filter(profesor=user.perfil_profesor).select_related("curso", "asignatura")
        cursos_ids = [a.curso_id for a in asignaciones]
        asignaturas_ids = [a.asignatura_id for a in asignaciones]
        alumnos = Alumno.objects.filter(curso_id__in=cursos_ids).select_related("usuario")
        asignaturas = Asignatura.objects.filter(id__in=asignaturas_ids)
    else:
        alumnos = Alumno.objects.all().select_related("usuario")
        asignaturas = Asignatura.objects.all()

    return render(request, "academicos/nota_form.html", {
        "alumnos": alumnos,
        "asignaturas": asignaturas,
        "tipos": Nota.Tipo.choices,
    })


@login_required
@rol_requerido("profesor", "admin")
def nota_carga_masiva(request):
    from apps.usuarios.models import Alumno
    from apps.gestion.models import Asignatura

    user = request.user
    asignatura_id = request.GET.get("asignatura")
    periodo = request.GET.get("periodo", "1er Trimestre")
    tipo = request.GET.get("tipo", "examen")

    if user.es_profesor():
        asignaciones = AsignacionCurso.objects.filter(profesor=user.perfil_profesor).select_related("curso", "asignatura")
        asignaturas_ids = [a.asignatura_id for a in asignaciones]
        asignaturas = Asignatura.objects.filter(id__in=asignaturas_ids)
    else:
        asignaturas = Asignatura.objects.all()

    alumnos = []
    if asignatura_id:
        asig = get_object_or_404(Asignatura, id=asignatura_id)
        cursos = AsignacionCurso.objects.filter(asignatura=asig).values_list("curso", flat=True)
        alumnos = Alumno.objects.filter(curso_id__in=cursos).select_related("usuario")

    if request.method == "POST":
        asig_id = request.POST.get("asignatura")
        periodo = request.POST.get("periodo", "1er Trimestre")
        tipo = request.POST.get("tipo", "examen")
        notas_creadas = 0
        for key, valor in request.POST.items():
            if key.startswith("nota_"):
                alumno_id = key.split("_")[1]
                if valor.strip():
                    Nota.objects.create(
                        alumno_id=alumno_id,
                        asignatura_id=asig_id,
                        profesor=user.perfil_profesor if user.es_profesor() else None,
                        valor=valor,
                        tipo=tipo,
                        periodo=periodo,
                        observacion=request.POST.get(f"obs_{alumno_id}", ""),
                    )
                    notas_creadas += 1
        messages.success(request, f"{notas_creadas} notas cargadas exitosamente")
        return redirect("nota_list")

    periodos = [
        "1er Trimestre", "2do Trimestre", "3er Trimestre",
        "1er Cuatrimestre", "2do Cuatrimestre",
    ]

    return render(request, "academicos/nota_carga_masiva.html", {
        "asignaturas": asignaturas,
        "alumnos": alumnos,
        "asignatura_id": asignatura_id,
        "periodo": periodo,
        "tipo": tipo,
        "tipos": Nota.Tipo.choices,
        "periodos": periodos,
    })


@login_required
@rol_requerido("admin")
def nota_exportar_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    q = request.GET.get("q", "")
    tipo = request.GET.get("tipo", "")
    periodo = request.GET.get("periodo", "")

    notas = Nota.objects.all().select_related("alumno__usuario", "asignatura", "profesor__usuario")
    if q:
        notas = notas.filter(
            Q(alumno__usuario__nombre__icontains=q) |
            Q(alumno__usuario__apellido__icontains=q) |
            Q(asignatura__nombre__icontains=q)
        )
    if tipo:
        notas = notas.filter(tipo=tipo)
    if periodo:
        notas = notas.filter(periodo__icontains=periodo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Notas"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Alumno", "Asignatura", "Nota", "Tipo", "Periodo", "Fecha", "Profesor", "Observación"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, nota in enumerate(notas, 2):
        ws.cell(row=i, column=1, value=str(nota.alumno))
        ws.cell(row=i, column=2, value=nota.asignatura.nombre)
        ws.cell(row=i, column=3, value=float(nota.valor))
        ws.cell(row=i, column=4, value=nota.get_tipo_display())
        ws.cell(row=i, column=5, value=nota.periodo)
        ws.cell(row=i, column=6, value=nota.fecha.isoformat())
        ws.cell(row=i, column=7, value=str(nota.profesor) if nota.profesor else "")
        ws.cell(row=i, column=8, value=nota.observacion)
        for col in range(1, 9):
            ws.cell(row=i, column=col).border = thin_border

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 25

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="notas_{date.today().isoformat()}.xlsx"'
    wb.save(response)
    return response


@login_required
def tarea_list(request):
    user = request.user
    q = request.GET.get("q", "")

    if user.es_alumno():
        tareas = Tarea.objects.filter(
            asignatura__curso=user.perfil_alumno.curso
        ).select_related("asignatura", "profesor__usuario")
    elif user.es_profesor():
        tareas = Tarea.objects.filter(profesor=user.perfil_profesor).select_related("asignatura", "profesor__usuario")
    else:
        tareas = Tarea.objects.all().select_related("asignatura", "profesor__usuario")

    if q:
        tareas = tareas.filter(
            Q(titulo__icontains=q) |
            Q(descripcion__icontains=q) |
            Q(asignatura__nombre__icontains=q)
        )

    paginator = Paginator(tareas, 12)
    page = paginator.get_page(request.GET.get("page"))

    return render(request, "academicos/tarea_list.html", {"tareas": page, "q": q})


@login_required
@rol_requerido("profesor", "admin")
def tarea_create(request):
    from apps.gestion.models import Asignatura as AsignaturaModel
    user = request.user
    if request.method == "POST":
        Tarea.objects.create(
            titulo=request.POST.get("titulo"),
            descripcion=request.POST.get("descripcion"),
            asignatura_id=request.POST.get("asignatura"),
            profesor=user.perfil_profesor if user.es_profesor() else None,
            fecha_entrega=request.POST.get("fecha_entrega"),
            archivo=request.FILES.get("archivo"),
        )
        messages.success(request, "Tarea creada exitosamente")
        return redirect("tarea_list")

    if user.es_profesor():
        asignaciones = AsignacionCurso.objects.filter(profesor=user.perfil_profesor).select_related("curso", "asignatura")
        asignaturas_ids = [a.asignatura_id for a in asignaciones]
        asignaturas = AsignaturaModel.objects.filter(id__in=asignaturas_ids)
    else:
        asignaturas = AsignaturaModel.objects.all()

    return render(request, "academicos/tarea_form.html", {"asignaturas": asignaturas})


@login_required
def entrega_list(request, tarea_pk=None):
    user = request.user
    estado = request.GET.get("estado", "")

    if tarea_pk:
        entregas = EntregaTarea.objects.filter(tarea_id=tarea_pk).select_related("alumno__usuario", "tarea")
    elif user.es_alumno():
        entregas = EntregaTarea.objects.filter(alumno=user.perfil_alumno).select_related("tarea")
    elif user.es_profesor():
        entregas = EntregaTarea.objects.filter(tarea__profesor=user.perfil_profesor).select_related("alumno__usuario", "tarea")
    else:
        entregas = EntregaTarea.objects.all().select_related("alumno__usuario", "tarea")

    if estado:
        entregas = entregas.filter(estado=estado)

    paginator = Paginator(entregas, 15)
    page = paginator.get_page(request.GET.get("page"))

    return render(request, "academicos/entrega_list.html", {
        "entregas": page,
        "estado_filtro": estado,
    })


@login_required
@rol_requerido("alumno")
def entrega_create(request):
    if request.method == "POST":
        tarea_id = request.POST.get("tarea")
        alumno = request.user.perfil_alumno
        EntregaTarea.objects.update_or_create(
            tarea_id=tarea_id,
            alumno=alumno,
            defaults={
                "archivo": request.FILES.get("archivo"),
                "estado": EntregaTarea.Estado.ENTREGADO,
                "observacion": request.POST.get("observacion", ""),
            },
        )
        messages.success(request, "Tarea entregada exitosamente")
        return redirect("tarea_list")

    tareas = Tarea.objects.filter(
        asignatura__curso=request.user.perfil_alumno.curso
    ).select_related("asignatura")
    return render(request, "academicos/entrega_form.html", {"tareas": tareas})


@login_required
@rol_requerido("profesor", "admin")
def entrega_calificar(request, pk):
    entrega = get_object_or_404(EntregaTarea, pk=pk)
    if request.method == "POST":
        entrega.nota = request.POST.get("nota")
        entrega.observacion = request.POST.get("observacion", "")
        entrega.estado = EntregaTarea.Estado.CALIFICADO
        entrega.save()
        messages.success(request, "Entrega calificada exitosamente")
        return redirect("entrega_list")
    return render(request, "academicos/entrega_calificar.html", {"entrega": entrega})


@login_required
@rol_requerido("profesor", "admin")
def asistencia_list(request):
    user = request.user
    fecha = request.GET.get("fecha", date.today().isoformat())
    asignatura_id = request.GET.get("asignatura")

    asistencias = Asistencia.objects.all().select_related("alumno__usuario", "asignatura")
    if asignatura_id:
        asistencias = asistencias.filter(asignatura_id=asignatura_id)
    if fecha:
        asistencias = asistencias.filter(fecha=fecha)

    from apps.gestion.models import Asignatura
    if user.es_profesor():
        asignaciones = AsignacionCurso.objects.filter(profesor=user.perfil_profesor)
        asignaturas = Asignatura.objects.filter(id__in=[a.asignatura_id for a in asignaciones])
    else:
        asignaturas = Asignatura.objects.all()

    return render(request, "academicos/asistencia_list.html", {
        "asistencias": asistencias,
        "asignaturas": asignaturas,
        "fecha": fecha,
        "asignatura_id": asignatura_id,
    })


@login_required
@rol_requerido("profesor", "admin")
def asistencia_tomar(request):
    from apps.usuarios.models import Alumno
    from apps.gestion.models import Asignatura

    user = request.user
    asignatura_id = request.GET.get("asignatura")
    fecha_str = request.GET.get("fecha", date.today().isoformat())

    if user.es_profesor():
        asignaciones = AsignacionCurso.objects.filter(profesor=user.perfil_profesor).select_related("curso", "asignatura")
        asignaturas_ids = [a.asignatura_id for a in asignaciones]
        asignaturas = Asignatura.objects.filter(id__in=asignaturas_ids)
    else:
        asignaturas = Asignatura.objects.all()

    alumnos = []
    if asignatura_id:
        asig = get_object_or_404(Asignatura, id=asignatura_id)
        cursos = AsignacionCurso.objects.filter(asignatura=asig).values_list("curso", flat=True)
        alumnos_qs = Alumno.objects.filter(curso_id__in=cursos).select_related("usuario")

        asistencias_existentes = Asistencia.objects.filter(
            asignatura_id=asignatura_id, fecha=fecha_str
        ).values_list("alumno_id", flat=True)

        alumnos = [
            {"alumno": a, "ya_registrado": a.id in asistencias_existentes}
            for a in alumnos_qs
        ]

    if request.method == "POST":
        asig_id = request.POST.get("asignatura")
        fecha = request.POST.get("fecha")
        contador = 0
        for key, value in request.POST.items():
            if key.startswith("presente_"):
                alumno_id = int(key.split("_")[1])
                presente = value == "true"
                Asistencia.objects.update_or_create(
                    alumno_id=alumno_id,
                    asignatura_id=asig_id,
                    fecha=fecha,
                    defaults={
                        "presente": presente,
                        "observacion": request.POST.get(f"obs_{alumno_id}", ""),
                        "registrado_por": user.perfil_profesor if user.es_profesor() else None,
                    },
                )
                contador += 1
        messages.success(request, f"Asistencia registrada para {contador} alumnos")
        return redirect("asistencia_list")

    return render(request, "academicos/asistencia_tomar.html", {
        "asignaturas": asignaturas,
        "alumnos": alumnos,
        "asignatura_id": asignatura_id,
        "fecha": fecha_str,
    })


@login_required
def calendario(request):
    user = request.user
    hoy = date.today()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))

    inicio_mes = date(anio, mes, 1)
    if mes == 12:
        fin_mes = date(anio + 1, 1, 1) - timedelta(days=1)
    else:
        fin_mes = date(anio, mes + 1, 1) - timedelta(days=1)

    if user.es_alumno():
        tareas = Tarea.objects.filter(
            asignatura__curso=user.perfil_alumno.curso,
            fecha_entrega__gte=inicio_mes,
            fecha_entrega__lte=fin_mes,
        ).select_related("asignatura")
        notas = Nota.objects.filter(
            alumno=user.perfil_alumno,
            fecha__gte=inicio_mes,
            fecha__lte=fin_mes,
        ).select_related("asignatura")
    elif user.es_profesor():
        tareas = Tarea.objects.filter(
            profesor=user.perfil_profesor,
            fecha_entrega__gte=inicio_mes,
            fecha_entrega__lte=fin_mes,
        ).select_related("asignatura")
        notas = Nota.objects.filter(
            profesor=user.perfil_profesor,
            fecha__gte=inicio_mes,
            fecha__lte=fin_mes,
        ).select_related("asignatura")
    else:
        tareas = Tarea.objects.filter(
            fecha_entrega__gte=inicio_mes, fecha_entrega__lte=fin_mes
        ).select_related("asignatura")
        notas = Nota.objects.filter(
            fecha__gte=inicio_mes, fecha__lte=fin_mes
        ).select_related("asignatura")

    eventos_por_dia = {}
    for t in tareas:
        d = t.fecha_entrega.day
        eventos_por_dia.setdefault(d, []).append({
            "titulo": t.titulo,
            "tipo": "tarea",
            "url": "/academicos/tareas/",
            "tooltip": f"Tarea: {t.titulo} - {t.asignatura.nombre}",
        })
    for n in notas:
        d = n.fecha.day
        eventos_por_dia.setdefault(d, []).append({
            "titulo": f"Nota: {n.valor}",
            "tipo": "nota",
            "url": "/academicos/notas/",
            "tooltip": f"Nota: {n.asignatura.nombre} - {n.valor}",
        })

    dias_del_mes = []
    dia_actual = inicio_mes
    while dia_actual <= fin_mes:
        dias_del_mes.append({
            "dia": dia_actual.day,
            "fecha": dia_actual,
            "eventos": eventos_por_dia.get(dia_actual.day, []),
        })
        dia_actual += timedelta(days=1)

    meses = [
        (1, "Enero"), (2, "Febrero"), (3, "Marzo"), (4, "Abril"),
        (5, "Mayo"), (6, "Junio"), (7, "Julio"), (8, "Agosto"),
        (9, "Septiembre"), (10, "Octubre"), (11, "Noviembre"), (12, "Diciembre"),
    ]

    mes_ant = 12 if mes == 1 else mes - 1
    anio_ant = anio - 1 if mes == 1 else anio
    mes_sig = 1 if mes == 12 else mes + 1
    anio_sig = anio + 1 if mes == 12 else anio

    return render(request, "academicos/calendario.html", {
        "dias": dias_del_mes,
        "mes": mes,
        "anio": anio,
        "meses": meses,
        "nombre_mes": dict(meses)[mes],
        "mes_ant": mes_ant,
        "anio_ant": anio_ant,
        "mes_sig": mes_sig,
        "anio_sig": anio_sig,
    })


@login_required
def comentario_list(request, tarea_pk):
    tarea = get_object_or_404(Tarea, pk=tarea_pk)
    comentarios = Comentario.objects.filter(tarea=tarea).select_related("usuario")
    if request.method == "POST":
        texto = request.POST.get("texto", "").strip()
        if texto:
            Comentario.objects.create(tarea=tarea, usuario=request.user, texto=texto)
            messages.success(request, "Comentario agregado")
        return redirect("comentario_list", tarea_pk=tarea_pk)
    return render(request, "academicos/comentario_list.html", {
        "tarea": tarea,
        "comentarios": comentarios,
    })


@login_required
def tarea_detail(request, pk):
    tarea = get_object_or_404(Tarea, pk=pk)
    comentarios = Comentario.objects.filter(tarea=tarea).select_related("usuario")
    return render(request, "academicos/tarea_detail.html", {
        "tarea": tarea,
        "comentarios": comentarios,
    })


@login_required
@rol_requerido("admin")
def boletin_pdf(request, alumno_id):
    from apps.usuarios.models import Alumno
    from django.template.loader import render_to_string
    from weasyprint import HTML

    alumno = get_object_or_404(Alumno.objects.select_related("usuario", "curso"), pk=alumno_id)
    notas = Nota.objects.filter(alumno=alumno).select_related("asignatura")

    html = render_to_string("academicos/boletin_pdf.html", {
        "alumno": alumno,
        "notas": notas,
    })
    pdf = HTML(string=html).write_pdf()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="boletin_{alumno.usuario.apellido}_{alumno.usuario.nombre}.pdf"'
    return response
